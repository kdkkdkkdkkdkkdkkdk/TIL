from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

for row in ws.iter_rows(min_row=2):
    # 번호, 영어, 수학
    if int(row[1].value) > 80:
        print(row[0].value, "번 학생은 영어 천재")

# 파이썬 코딩 무료 강의 (활용편4) - 업무자동화(RPA), 이제는 일하는 척(?)만 하세요
# 1:04:52
